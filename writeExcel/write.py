import openpyxl
import urllib.request
import datetime
import time
import os
import re

filePath = "F:/E/git/python-learn/writeExcel/download/" 
#换成自己的下载目录地址
wb = openpyxl.load_workbook('file/allhref.xlsx')
#换成自己的exal目录

sheets = wb.sheetnames
# print(sheets, type(sheets)) 
import dowloadXunlei
import common



def getHref(ws, sheet):
    # print(ws['A']) A一竖列
    #循环B数列
    overNum = 0
    i = 0
    for href in  ws['B']: 
        name = (ws['A'][i].value)
        isOver = (ws['D'][i].value)
        i = i + 1
        if href.value != 'href' and href.value: 
            if not isOver: 
                common.WriteTime(i, common.getCurrentTime(), 'C')
                getFileType(href.value, name, i)
                overNum = overNum + 1
    if(overNum == 0):
        print('没有新的任务')
             

def getFileType(url, name, i): 
    if(re.search("jpg",url)):
        dowloadPic(url, name, i)
    elif(re.search("rmvb",url)):
       dowloadXunlei.download(url, name, i)
    elif(re.search("mkv",url)):
       dowloadXunlei.download(url, name, i) 
    elif(re.search("mp4",url)):
       dowloadXunlei.download(url, name, i) 





def dowloadPic(url, fileName, i):
        response = urllib.request.urlopen(url)
        data = response.read()
        t = int(time.time() * 1000)
        os.mkdir("download/"+fileName)
        name = filePath +fileName+"/"+ '%d'%t+".jpg"
        print(name)
        with open(name, 'wb') as code:
            code.write(data)
            print('finish')
            WriteTime(i, common.getCurrentTime(), 'D')


for sheet in sheets:        # 循环表
   if(sheet == 'Sheet1'): getHref(wb[sheet], sheet)
        # wb[sheet]




